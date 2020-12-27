import openpyxl

wb = openpyxl.Workbook()
 
sheet = wb.get_sheet_by_name('Sheet') # initializarea foii de lucru

sheet['A1'] = 42
sheet['A2'] = 'Hello'

sheet2 = wb.create_sheet() # create a new sheet
wb.create_sheet(index = 0, title = 'My other sheet') # create a sheet with proprieties
sheet2.title = 'Cam nasol asa' # rename sheet name

wb.save('example1.xlsx') # it saves in the curent working directory
