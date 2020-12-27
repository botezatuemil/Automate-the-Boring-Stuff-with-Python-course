import openpyxl

workbook = openpyxl.load_workbook('example.xlsx')
print(type(workbook))

sheet = workbook.get_sheet_by_name('Sheet1')
print(type(sheet))

print(workbook.get_sheet_names()) # return a list with name of sheets

cell = sheet['A1']
print(str(cell.value)) # return the value of cell

print(sheet['B1'].value) # same thing as above

print(sheet.cell(row = 1, column = 2).value) # same thing as above

for i in range(1, 8):
    print(i, sheet.cell(row = i, column = 2).value) # print the entire column

print('\n')
for i in range(1, 4):
    print(i, sheet.cell(row = 1, column = i).value) # print the entire row

for i in range(1, 8):
    for j in range(1, 4):
        print(sheet.cell(row = i, column = j).value, end = ' ') # print the entire table
    print('\n')
